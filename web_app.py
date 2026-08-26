import os
import re
import io
import json
import base64
import datetime
import requests
from PIL import Image
import openpyxl
import streamlit as st

# ==========================================
# 🌟 Page Configuration & Title
# ==========================================
st.set_page_config(
    page_title="User Access Automation | MGC-Asia",
    page_icon="🔐",
    layout="wide",
    initial_sidebar_state="expanded"
)

MASTER_DATA_FILE = "master_data.json"

DEFAULT_MASTER_DATA = {
    "companies": [
        {"Company": "Belfort Automobile (Thailand) Co., Ltd.", "BU": "Belfort"},
        {"Company": "Gaydon Motor Sales and Services Co., Ltd.", "BU": "Gaydon"},
        {"Company": "Goodwood Autowork Co., Ltd.", "BU": "Goodwood"},
        {"Company": "Howden Maxi Insurance Broker Co., Ltd.", "BU": "Howden Maxi"},
        {"Company": "i24 Co., Ltd.", "BU": "I24"},
        {"Company": "Lion Automobile Co., Ltd.", "BU": "Lion"},
        {"Company": "Master Car Rental Co., Ltd.", "BU": "MCR"},
        {"Company": "Master Driver & Services (Thailand) Co., Ltd.", "BU": "MDS"},
        {"Company": "Master Group Corporation (Laos) Co., Ltd.", "BU": "MGC Laos"},
        {"Company": "Master Motor Services (Thailand) Co., Ltd.", "BU": "MMS"},
        {"Company": "MGC Aviation and Charter Service (Asia) Co., Ltd.", "BU": "MGC Aviation"},
        {"Company": "MGC Marine & Charter (Asia) Co., Ltd.", "BU": "MGC Marine"},
        {"Company": "Millennium Auto Group Co., Ltd.", "BU": "BMW"},
        {"Company": "Millennium Group Corporation (ASIA) Co., Ltd.", "BU": "MGC"},
        {"Company": "Modena Motorwork Co., Ltd.", "BU": "Modena"},
        {"Company": "Summit Honda Automobile Co., Ltd.", "BU": "Summit Honda"},
        {"Company": "US Motorbike Co., Ltd.", "BU": "Harley"},
        {"Company": "X Mobility Plus Co.,Ltd", "BU": "XP"},
        {"Company": "X Mobility Thailand", "BU": "X Mobility"},
        {"Company": "Ze Mobility Plus co., ltd", "BU": "Ze Mobility"}
    ],
    "branches": [
        "กรุงเทพมหานคร", "สำนักงานใหญ่", "พระราม 3", "พระราม 4", "ลาดพร้าว", "รามคำแหง",
        "อุดรธานี", "ภูเก็ต", "หาดใหญ่", "อุบลราชธานี", 
        "สุราษฎร์ธานี", "พัทยา", "เชียงใหม่", "สยามพารากอน", "ไอคอนสยาม"
    ],
    "operators": [
        "nawarutte.non@i24.co.th",
        "pawitporn.sae@i24.co.th"
    ],
    "templates": {
        "nawarutte.non@i24.co.th": {
            "VSM": {
                "subject": "ข้อมูลการเข้าระบบ VSM",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งานโดย User & Password ตามด้านล่างนี้ครับ"
            },
            "E-Travelling": {
                "subject": "ข้อมูลการเข้าระบบ E-Travelling",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งานโดย User & Password ตามด้านล่างนี้ครับ"
            },
            "Forma": {
                "subject": "ข้อมูลการเข้าระบบ Forma",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งาน Forma โดย User & Password ตามด้านล่างนี้ครับ"
            },
            "Red plate": {
                "subject": "ข้อมูลการเข้าระบบ Red plate",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งาน Red plate โดย User & Password ตามด้านล่างนี้ครับ"
            },
            "Pandora": {
                "subject": "ข้อมูลการเข้าระบบ Pandora",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งาน Pandora โดย User & Password ตามด้านล่างนี้ครับ"
            }
        },
        "pawitporn.sae@i24.co.th": {
            "VSM": {
                "subject": "ข้อมูลการเข้าระบบ VSM",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งานโดย User & Password ตามด้านล่างนี้ครับ"
            },
            "E-Travelling": {
                "subject": "ข้อมูลการเข้าระบบ E-Travelling",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งานโดย User & Password ตามด้านล่างนี้ครับ"
            },
            "Forma": {
                "subject": "ข้อมูลการเข้าระบบ Forma",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งาน Forma โดย User & Password ตามด้านล่างนี้ครับ"
            },
            "Red plate": {
                "subject": "ข้อมูลการเข้าระบบ Red plate",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งาน Red plate โดย User & Password ตามด้านล่างนี้ครับ"
            },
            "Pandora": {
                "subject": "ข้อมูลการเข้าระบบ Pandora",
                "greeting": "เรียน ผู้ใช้งานระบบ",
                "intro": "ให้เข้าใช้งาน Pandora โดย User & Password ตามด้านล่างนี้ครับ"
            }
        }
    }
}

def load_master_data():
    """โหลดข้อมูล Master Data จากไฟล์ JSON หรือใช้ค่าเริ่มต้น"""
    if os.path.exists(MASTER_DATA_FILE):
        try:
            with open(MASTER_DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if not data.get("companies"):
                    data["companies"] = DEFAULT_MASTER_DATA["companies"]
                if not data.get("branches"):
                    data["branches"] = DEFAULT_MASTER_DATA["branches"]
                if not data.get("operators"):
                    data["operators"] = DEFAULT_MASTER_DATA["operators"]
                if not data.get("templates"):
                    data["templates"] = DEFAULT_MASTER_DATA["templates"]
                return data
        except Exception:
            pass
    return json.loads(json.dumps(DEFAULT_MASTER_DATA))

def save_master_data(data):
    """บันทึก Master Data ลงไฟล์ JSON"""
    try:
        with open(MASTER_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        st.error(f"เกิดข้อผิดพลาดในการบันทึก Master Data: {e}")
        return False

def get_operator_template(master_data, operator, app_name):
    """ดึง Template อีเมลเฉพาะของ User และ App นั้นๆ"""
    templates = master_data.get("templates", {})
    user_tpl = templates.get(operator, {})
    if not user_tpl:
        user_tpl = templates.get("nawarutte.non@i24.co.th", {})
    
    app_tpl = user_tpl.get(app_name)
    if not app_tpl:
        default_tpl = DEFAULT_MASTER_DATA["templates"]["nawarutte.non@i24.co.th"].get(app_name, {
            "subject": f"ข้อมูลการเข้าระบบ {app_name}",
            "greeting": "เรียน ผู้ใช้งานระบบ",
            "intro": f"ให้เข้าใช้งาน {app_name} โดย User & Password ตามด้านล่างนี้ครับ"
        })
        return default_tpl
    return app_tpl

def map_branch_name(branch_str):
    if not branch_str:
        return ""
    b_lower = str(branch_str).lower().strip()
    if "ladprao" in b_lower or "ลาดพร้าว" in b_lower:
        return "ลาดพร้าว"
    elif "rama 3" in b_lower or "พระราม 3" in b_lower or "rama3" in b_lower:
        return "พระราม 3"
    elif "rama 4" in b_lower or "พระราม 4" in b_lower or "rama4" in b_lower:
        return "พระราม 4"
    elif "phuket" in b_lower or "ภูเก็ต" in b_lower:
        return "ภูเก็ต"
    elif "udon" in b_lower or "อุดร" in b_lower:
        return "อุดรธานี"
    elif "hatyai" in b_lower or "hat yai" in b_lower or "หาดใหญ่" in b_lower:
        return "หาดใหญ่"
    elif "ubon" in b_lower or "อุบล" in b_lower:
        return "อุบลราชธานี"
    elif "surat" in b_lower or "สุราษฎร์" in b_lower:
        return "สุราษฎร์ธานี"
    elif "pattaya" in b_lower or "พัทยา" in b_lower:
        return "พัทยา"
    elif "chiangmai" in b_lower or "chiang mai" in b_lower or "เชียงใหม่" in b_lower:
        return "เชียงใหม่"
    elif "paragon" in b_lower or "พารากอน" in b_lower:
        return "สยามพารากอน"
    elif "icon" in b_lower or "ไอคอน" in b_lower:
        return "ไอคอนสยาม"
    elif "head office" in b_lower or "สำนักงานใหญ่" in b_lower or "head" in b_lower:
        return "สำนักงานใหญ่"
    elif "bangkok" in b_lower or "กรุงเทพ" in b_lower or "bkk" in b_lower:
        return "กรุงเทพมหานคร"
    elif "ramkhamhaeng" in b_lower or "รามคำแหง" in b_lower:
        return "รามคำแหง"
    return branch_str

COMPANY_BU_MAPPING = {
    "mastercarrental.com": {"Company": "Master Car Rental Co., Ltd.", "BU": "MCR"},
    "i24.co.th": {"Company": "i24 Co., Ltd.", "BU": "I24"},
    "bmw-millenniumauto.com": {"Company": "Millennium Auto Group Co., Ltd.", "BU": "BMW"},
    "millenniumauto.co.th": {"Company": "Millennium Auto Group Co., Ltd.", "BU": "BMW"},
    "usmotorbike.com": {"Company": "US Motorbike Co., Ltd.", "BU": "Harley"},
    "mgc-asia.com": {"Company": "", "BU": ""}
}

# ==========================================
# 🛠️ Helper & Password Functions
# ==========================================
def generate_vsm_password(email):
    if not email:
        return "p@ssw0rd***"
    local_part = email.split('@')[0] if '@' in email else email
    letters = re.findall(r'[a-zA-Z]', local_part)
    if len(letters) >= 3:
        prefix = "".join(letters[:3]).lower()
    else:
        prefix = "".join(letters).lower().ljust(3, 'x')
    return f"p@ssw0rd{prefix}"

def generate_forma_password(user_login):
    if not user_login:
        return "p@ssw0rd***"
    letters = re.findall(r'[a-zA-Z]', user_login)
    if len(letters) >= 3:
        prefix = "".join(letters[:3]).lower()
    else:
        prefix = "".join(letters).lower().ljust(3, 'x')
    return f"p@ssw0rd{prefix}"

def get_default_password_for_app(app_name, user_id, email, extracted_pwd=""):
    if app_name == "Pandora" and extracted_pwd:
        return extracted_pwd
    elif app_name == "VSM":
        return generate_vsm_password(email)
    elif app_name in ["Forma", "Pandora"]:
        return generate_forma_password(user_id)
    elif app_name == "Red plate":
        return "Init123456"
    else:
        return user_id

def build_email_body(app_name, user_id, password_str, link_str, custom_template=None):
    if custom_template:
        greeting = custom_template.get("greeting", "เรียน ผู้ใช้งานระบบ")
        intro = custom_template.get("intro", f"ให้เข้าใช้งานโดย User & Password ตามด้านล่างนี้ครับ")
    elif app_name in ["Forma", "Pandora"]:
        greeting = "เรียน ผู้ใช้งานระบบ"
        intro = f"ให้เข้าใช้งาน {app_name} โดย User & Password ตามด้านล่างนี้ครับ"
    elif app_name == "Red plate":
        greeting = "เรียน ผู้ใช้งานระบบ"
        intro = "ให้เข้าใช้งาน Red plate โดย User & Password ตามด้านล่างนี้ครับ"
    else:
        greeting = "เรียน ผู้ใช้งานระบบ"
        intro = "ให้เข้าใช้งานโดย User & Password ตามด้านล่างนี้ครับ"

    text_lines = [
        greeting,
        intro,
        f"User: {user_id}",
        f"Password: {password_str}"
    ]
    if link_str:
        text_lines.append(f"Link: {link_str}")
    plain_text = "\n".join(text_lines)

    html_parts = [
        f"<p>{greeting}</p>",
        f"<p>{intro}</p>",
        f"<p><b>User:</b> {user_id}<br><b>Password:</b> {password_str}"
    ]
    if link_str:
        html_parts.append(f"<br><b>Link:</b> <a href=\"{link_str}\">{link_str}</a>")
    html_parts.append("</p>")
    html_body = "".join(html_parts)

    return plain_text, html_body

def derive_company_and_bu(company_input, email_input, master_companies=None):
    email_lower = (email_input or "").lower()
    comp_lower = (company_input or "").lower()
    
    # 1. ตรวจสอบกับ Master Data Companies
    if master_companies and comp_lower:
        for c_entry in master_companies:
            c_name = c_entry.get("Company", "")
            c_bu = c_entry.get("BU", "")
            if c_name.lower() == comp_lower or comp_lower in c_name.lower():
                return c_name, c_bu

    if comp_lower:
        if "master car" in comp_lower or "mcr" in comp_lower or "mastercar" in comp_lower:
            return "Master Car Rental Co., Ltd.", "MCR"
        elif "millennium auto" in comp_lower or "bmw" in comp_lower:
            return "Millennium Auto Group Co., Ltd.", "BMW"
        elif "us motorbike" in comp_lower or "harley" in comp_lower:
            return "US Motorbike Co., Ltd.", "Harley"
        elif "i24" in comp_lower:
            return "i24 Co., Ltd.", "I24"
        elif "belfort" in comp_lower or "jeep" in comp_lower or "peugeot" in comp_lower:
            return "Belfort Automobile (Thailand) Co., Ltd.", "Belfort"
        elif "gaydon" in comp_lower or "aston martin" in comp_lower:
            return "Gaydon Motor Sales and Services Co., Ltd.", "Gaydon"
        elif "goodwood" in comp_lower or "rolls" in comp_lower:
            return "Goodwood Autowork Co., Ltd.", "Goodwood"
        elif "howden" in comp_lower or "maxi" in comp_lower:
            return "Howden Maxi Insurance Broker Co., Ltd.", "Howden Maxi"
        elif "lion automobile" in comp_lower or "lion" in comp_lower:
            return "Lion Automobile Co., Ltd.", "Lion"
        elif "master driver" in comp_lower or "mds" in comp_lower:
            return "Master Driver & Services (Thailand) Co., Ltd.", "MDS"
        elif "laos" in comp_lower:
            return "Master Group Corporation (Laos) Co., Ltd.", "MGC Laos"
        elif "master motor" in comp_lower or "mms" in comp_lower:
            return "Master Motor Services (Thailand) Co., Ltd.", "MMS"
        elif "aviation" in comp_lower:
            return "MGC Aviation and Charter Service (Asia) Co., Ltd.", "MGC Aviation"
        elif "marine" in comp_lower or "charter" in comp_lower:
            return "MGC Marine & Charter (Asia) Co., Ltd.", "MGC Marine"
        elif "modena" in comp_lower or "maserati" in comp_lower:
            return "Modena Motorwork Co., Ltd.", "Modena"
        elif "summit honda" in comp_lower or "summit" in comp_lower or "honda" in comp_lower:
            return "Summit Honda Automobile Co., Ltd.", "Summit Honda"
        elif "x mobility plus" in comp_lower or "x-mobility plus" in comp_lower:
            return "X Mobility Plus Co.,Ltd", "XP"
        elif "x mobility" in comp_lower or "x-mobility" in comp_lower:
            return "X Mobility Thailand", "X Mobility"
        elif "ze mobility" in comp_lower or "ze-mobility" in comp_lower or "zeekr" in comp_lower:
            return "Ze Mobility Plus co., ltd", "Ze Mobility"
        elif "millennium group" in comp_lower or "mgc" in comp_lower:
            return "Millennium Group Corporation (ASIA) Co., Ltd.", "MGC"
            
    if '@' in email_lower:
        domain = email_lower.split('@')[-1].strip()
        for dom_key, info in COMPANY_BU_MAPPING.items():
            if dom_key in domain:
                return info["Company"], info["BU"]
                
    return company_input if company_input else "", ""

def fix_email_domain(email_str):
    if not email_str or '@' not in email_str:
        return email_str
    local_part, domain = email_str.split('@', 1)
    domain_lower = domain.lower()
    if 'mastercar' in domain_lower:
        domain = 'mastercarrental.com'
    elif 'i2a' in domain_lower or 'i24' in domain_lower:
        domain = 'i24.co.th'
    elif 'millennium' in domain_lower or 'bmw' in domain_lower:
        domain = 'bmw-millenniumauto.com'
    elif 'mgc' in domain_lower or 'asia' in domain_lower:
        domain = 'mgc-asia.com'
    return f"{local_part}@{domain}"

# ==========================================
# 🤖 Vision AI Extraction
# ==========================================
def extract_with_gemini_vision(img, api_key):
    clean_key = api_key.strip()
    if not clean_key:
        return None, "กรุณากรอก Gemini API Key ในแถบตั้งค่าด้านซ้าย"
        
    prompt = """คุณคือผู้เชี่ยวชาญการอ่านข้อมูลฟอร์มระบบ VSM, E-Travelling, Forma, Red plate (Update User Info), และระบบ Pandora ของ MGC-Asia
จงวิเคราะห์ภาพแคปเจอร์นี้ และตอบเป็น JSON บริสุทธิ์เท่านั้น (ไม่ต้องมี markdown backticks) ในรูปแบบดังนี้:
{
  "Application": "Pandora" หรือ "Red plate" หรือ "Forma" หรือ "VSM" หรือ "E-Travelling",
  "App user ID": "รหัสผู้ใช้ สำหรับ Pandora หรือ Username สำหรับ Red plate หรือ Login สำหรับ Forma หรือ User Login สำหรับ VSM หรือ Employee Code สำหรับ E-Travelling",
  "Password": "รหัสผ่าน สำหรับ Pandora (เช่น p@ssw0rdcha) (ถ้าไม่มีให้เป็นว่างเปล่า \"\")",
  "Full Name Eng": "ชื่อ-นามสกุลภาษาอังกฤษ ถ้าเป็นภาษาอังกฤษล้วน (ตัดคำนำหน้าออก) ถ้าไม่มีให้เป็น \"\"",
  "Full Name Thai": "ชื่อ-นามสกุลภาษาไทย สำหรับ Red plate ให้นำ First Name เว้นวรรค Last Name (หรือ Display Name เช่น กัณฑิชา ลมหวล) มาใส่ช่องนี้ หรือชื่อไทยสำหรับระบบอื่น (ตัดคำนำหน้าออก) ถ้าไม่มีให้เป็น \"\"",
  "Email": "Email (ถ้าไม่มีในภาพให้เป็นว่างเปล่า \"\")",
  "Position": "กลุ่มผู้ใช้ สำหรับ Pandora (เช่น Accounting + Price) หรือ Role สำหรับ Red plate หรือ Position/User Type สำหรับระบบอื่น",
  "Company": "ชื่อบริษัท (ถ้าไม่มีในภาพให้เป็นว่างเปล่า \"\")",
  "Branch": "ชื่อสาขาจากภาพ เช่น Ladprao, Head Office, พระราม 3, ลาดพร้าว, รามคำแหง (ถ้าไม่มีในภาพให้เป็น \"\")"
}"""

    last_err = ""

    # 1. ลองใช้ google.genai SDK
    try:
        from google import genai
        from google.genai import types
        
        client = genai.Client(api_key=clean_key)
        models_to_try = ['gemini-3.5-flash', 'gemini-flash-latest', 'gemini-3.6-flash']
        for m_name in models_to_try:
            try:
                response = client.models.generate_content(
                    model=m_name,
                    contents=[prompt, img],
                    config=types.GenerateContentConfig(response_mime_type="application/json")
                )
                raw_text = response.text.strip()
                raw_text = re.sub(r'^```json\s*', '', raw_text, flags=re.IGNORECASE)
                raw_text = re.sub(r'```$', '', raw_text).strip()
                parsed_data = json.loads(raw_text)
                
                app_name = parsed_data.get("Application", "E-Travelling")
                user_id = str(parsed_data.get("App user ID", "")).strip()
                if app_name == "VSM" and user_id:
                    user_id = user_id.replace('.', '')
                    if len(user_id) > 3:
                        user_id = user_id[:-3] + '.' + user_id[-3:]
                    parsed_data["App user ID"] = user_id
                elif app_name == "Red plate":
                    pos = str(parsed_data.get("Position", "")).strip()
                    if "sales person" in pos.lower():
                        parsed_data["Position"] = "Sales Consultant"
                    nth = str(parsed_data.get("Full Name Thai", "")).strip()
                    neng = str(parsed_data.get("Full Name Eng", "")).strip()
                    if not nth and any('\u0e00' <= c <= '\u0e7f' for c in neng):
                        parsed_data["Full Name Thai"] = neng
                        parsed_data["Full Name Eng"] = ""
                    elif nth and not any('\u0e00' <= c <= '\u0e7f' for c in nth) and not neng:
                        parsed_data["Full Name Eng"] = nth
                        parsed_data["Full Name Thai"] = ""
                
                branch = str(parsed_data.get("Branch", "")).strip()
                if branch:
                    parsed_data["Branch"] = map_branch_name(branch)

                email = str(parsed_data.get("Email", "")).strip()
                if email:
                    parsed_data["Email"] = fix_email_domain(email)
                return parsed_data, ""
            except Exception as ex:
                last_err = str(ex)
                continue
    except Exception as g_err:
        last_err = str(g_err)

    # 2. REST API Fallback (HTTP Direct)
    try:
        buffered = io.BytesIO()
        img.convert('RGB').save(buffered, format="JPEG", quality=85)
        img_b64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        
        endpoints = [
            ("v1beta", "gemini-3.6-flash"),
            ("v1beta", "gemini-3.5-flash"),
            ("v1beta", "gemini-flash-latest")
        ]
        
        for api_ver, model_name in endpoints:
            url = f"https://generativelanguage.googleapis.com/{api_ver}/models/{model_name}:generateContent?key={clean_key}"
            headers = {"Content-Type": "application/json", "x-goog-api-key": clean_key}
                
            payload = {
                "contents": [
                    {
                        "parts": [
                            {"text": prompt},
                            {
                                "inline_data": {
                                    "mime_type": "image/jpeg",
                                    "data": img_b64
                                }
                            }
                        ]
                    }
                ],
                "generationConfig": {
                    "response_mime_type": "application/json"
                }
            }
            
            res = requests.post(url, json=payload, headers=headers, timeout=25)
            if res.status_code == 200:
                result_json = res.json()
                raw_text = result_json['candidates'][0]['content']['parts'][0]['text'].strip()
                raw_text = re.sub(r'^```json\s*', '', raw_text, flags=re.IGNORECASE)
                raw_text = re.sub(r'```$', '', raw_text).strip()
                
                parsed_data = json.loads(raw_text)
                
                app_name = parsed_data.get("Application", "E-Travelling")
                user_id = str(parsed_data.get("App user ID", "")).strip()
                if app_name == "VSM" and user_id:
                    user_id = user_id.replace('.', '')
                    if len(user_id) > 3:
                        user_id = user_id[:-3] + '.' + user_id[-3:]
                elif app_name == "Red plate":
                    pos = str(parsed_data.get("Position", "")).strip()
                    if "sales person" in pos.lower():
                        parsed_data["Position"] = "Sales Consultant"
                    nth = str(parsed_data.get("Full Name Thai", "")).strip()
                    neng = str(parsed_data.get("Full Name Eng", "")).strip()
                    if not nth and any('\u0e00' <= c <= '\u0e7f' for c in neng):
                        parsed_data["Full Name Thai"] = neng
                        parsed_data["Full Name Eng"] = ""
                branch = str(parsed_data.get("Branch", "")).strip()
                if branch:
                    parsed_data["Branch"] = map_branch_name(branch)
                    
                email = str(parsed_data.get("Email", "")).strip()
                if email:
                    parsed_data["Email"] = fix_email_domain(email)
                    
                return parsed_data, ""
            else:
                try:
                    err_msg = res.json().get('error', {}).get('message', res.text[:120])
                except Exception:
                    err_msg = res.text[:120]
                last_err = f"Google API Status {res.status_code}: {err_msg}"
    except Exception as ex:
        last_err = str(ex)

    if "invalid authentication" in last_err.lower() or "401" in last_err or "api key not valid" in last_err.lower():
        hint = " (หมายเหตุ: คีย์จาก Google AI Studio จะขึ้นต้นด้วย AIzaSy... กรุณากดปุ่ม 'รับ API Key ฟรี' ที่ aistudio.google.com/app/apikey เพื่อรับคีย์ที่ถูกต้อง)"
    else:
        hint = ""

    return None, f"ไม่สามารถสกัดข้อมูลจากภาพได้: {last_err}{hint}"

def process_mapping(extracted_data, master_companies=None):
    email = extracted_data.get("Email", "")
    comp_input = extracted_data.get("Company", "")
    company, bu = derive_company_and_bu(comp_input, email, master_companies=master_companies)

    branch_eng = extracted_data.get("Branch", "")
    now = datetime.datetime.now()
    current_date = datetime.date(now.year, now.month, now.day)

    return {
        "NO": 1,
        "Application": extracted_data.get("Application", "E-Travelling"),
        "App user ID": extracted_data.get("App user ID", ""),
        "Full Name Eng": extracted_data.get("Full Name Eng", ""),
        "Full Name Thai": extracted_data.get("Full Name Thai", ""),
        "Email": email,
        "Position": extracted_data.get("Position", ""),
        "Company": company,
        "BU": bu,
        "Area": extracted_data.get("Area", ""),
        "Role ID": extracted_data.get("Role ID", ""),
        "Branch": branch_eng,
        "Team": extracted_data.get("Team", ""),
        "Sub-Team": extracted_data.get("Sub-Team", ""),
        "x": "",
        "Type": "",
        "Parent Name (Manager)": extracted_data.get("Parent Name (Manager)", ""),
        "Status": "Active", 
        "Create date": current_date,
        "Disable date": "",
        "Remark": ""
    }

def export_to_excel(excel_row, excel_filename):
    if not excel_filename:
        excel_filename = "Template Column Excel Summary User.xlsx"
    try:
        if os.path.exists(excel_filename):
            wb = openpyxl.load_workbook(excel_filename)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = list(excel_row.keys())
            ws.append(headers)

        header_row = [cell.value for cell in ws[1]]
        max_no = 1
        if "NO" in header_row:
            no_idx = header_row.index("NO") + 1
            nos = []
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=no_idx).value
                if val is not None and str(val).isdigit():
                    nos.append(int(val))
            if nos:
                max_no = max(nos) + 1

        excel_row["NO"] = max_no
        row_data = [excel_row.get(col, "") for col in header_row]
        ws.append(row_data)
        wb.save(excel_filename)
        return True, f"บันทึกข้อมูลลง Excel ({os.path.basename(excel_filename)}) ลำดับที่ {max_no} สำเร็จ!"
    except Exception as e:
        return False, f"เกิดข้อผิดพลาดในการบันทึก Excel: {e}"

def trigger_power_automate_webhook(webhook_url, raw_data, mapped_data, custom_username=None, custom_password=None, send_email=True, operator="nawarutte.non@i24.co.th", custom_template=None):
    if not webhook_url:
        return False, "ไม่ได้ระบุ Webhook URL"
    try:
        app_name = raw_data.get("Application", "E-Travelling")
        email_to = raw_data.get("Email", "")
        user_id = custom_username if custom_username else raw_data.get("App user ID", "")
        
        if custom_password:
            password_str = custom_password
        elif app_name == "VSM":
            password_str = generate_vsm_password(email_to)
        elif app_name in ["Forma", "Pandora"]:
            password_str = generate_forma_password(user_id)
        elif app_name == "Red plate":
            password_str = "Init123456"
        else:
            password_str = user_id
            
        link_str = ""
        if app_name == "VSM":
            link_str = "https://vsm.mgc-asia.com/Pages/Home.aspx"
        elif app_name == "E-Travelling":
            link_str = "http://travelling.mgc-asia.com/"
        elif app_name == "Red plate":
            link_str = "https://redplate-frontend.azurewebsites.net/signin/?redirect_url=%2Freport%2F%3Ftype%3Dred-plate-transaction"

        plain_body, html_body = build_email_body(app_name, user_id, password_str, link_str, custom_template=custom_template)
        email_subject = custom_template.get("subject", f"ข้อมูลการเข้าระบบ {app_name}") if custom_template else f"ข้อมูลการเข้าระบบ {app_name}"

        payload = {
            "Application": app_name,
            "Email": email_to,
            "AppUserID": user_id,
            "Password": password_str,
            "Link": link_str,
            "LinkText": f"Link: {link_str}" if link_str else "",
            "EmailBody": plain_body,
            "EmailBodyHtml": html_body,
            "Subject": email_subject,
            "FullNameThai": raw_data.get("Full Name Thai", ""),
            "FullNameEng": raw_data.get("Full Name Eng", ""),
            "Position": raw_data.get("Position", ""),
            "Company": raw_data.get("Company", ""),
            "BU": mapped_data.get("BU", ""),
            "Area": mapped_data.get("Area", ""),
            "RoleID": mapped_data.get("Role ID", ""),
            "Branch": mapped_data.get("Branch", ""),
            "Team": mapped_data.get("Team", ""),
            "SubTeam": mapped_data.get("Sub-Team", ""),
            "ParentNameManager": mapped_data.get("Parent Name (Manager)", ""),
            "Status": "Active",
            "CreateDate": datetime.date.today().strftime('%Y-%m-%d'),
            "SendEmail": send_email,
            "SendEmailText": "true" if send_email else "false",
            "Operator": operator
        }
        res = requests.post(webhook_url, json=payload, headers={"Content-Type": "application/json"}, timeout=15)
        if res.status_code in [200, 202]:
            return True, "ส่งสัญญาณ Power Automate Webhook สำเร็จ!"
        return False, f"Power Automate Webhook ตอบกลับ Status: {res.status_code}"
    except Exception as e:
        return False, f"ข้อผิดพลาด Webhook: {e}"

WEB_CONFIG_FILE = "web_config.json"

def load_web_config():
    default_excel = r"C:\Users\Nawarutte.Non\OneDrive - Millennium Group Corporation (Asia) Public Company Limited\Automan\Template Column Excel Summary User.xlsx"
    if not os.path.exists(DEFAULT_EXCEL_PATH if 'DEFAULT_EXCEL_PATH' in locals() else default_excel):
        default_excel = "Template Column Excel Summary User.xlsx"
    
    cfg = {
        "gemini_api_key": os.environ.get("GEMINI_API_KEY", ""),
        "excel_path": default_excel,
        "webhook_url": os.environ.get("WEBHOOK_URL", "https://default6345207c7bd249f1920ea5aa88e4c1.c0.environment.api.powerplatform.com:443/powerautomate/automations/direct/cu/06/workflows/c8f4931f9e5646a08603ea1e9a63c307/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=Ue9CmEeB2GiJGWDyDWsCFpE7QcPcSYwXnKcXutGqRp0")
    }
    
    for fname in [WEB_CONFIG_FILE, "app_config.json"]:
        if os.path.exists(fname):
            try:
                with open(fname, "r", encoding="utf-8") as f:
                    saved = json.load(f)
                    if saved.get("gemini_api_key"):
                        cfg["gemini_api_key"] = saved["gemini_api_key"]
                    if saved.get("excel_path"):
                        cfg["excel_path"] = saved["excel_path"]
                    if saved.get("webhook_url"):
                        cfg["webhook_url"] = saved["webhook_url"]
            except Exception:
                pass
    return cfg

def save_web_config(api_k, exc_p, wh_u):
    try:
        data = {
            "gemini_api_key": api_k.strip(),
            "excel_path": exc_p.strip(),
            "webhook_url": wh_u.strip()
        }
        with open(WEB_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

CURRENT_WEB_CFG = load_web_config()

# ==========================================
# 🖥️ Sidebar & Config Setup
# ==========================================
with st.sidebar:
    st.image("https://img.icons8.com/color/96/000000/lock--v1.png", width=64)
    st.title("⚙️ ตั้งค่าระบบ")
    
    gemini_key = st.text_input(
        "🔑 Gemini Vision API Key (แม่นยำ 100% ฟรี)",
        value=CURRENT_WEB_CFG.get("gemini_api_key", ""),
        type="default",
        placeholder="วาง API Key ที่นี่ (ระบบจะจดจำคีย์ล่าสุดให้อัตโนมัติ)",
        help="รับ API Key ฟรีได้จาก https://aistudio.google.com/app/apikey"
    )
    
    excel_path = st.text_input(
        "📂 ไฟล์ Excel Summary Target",
        value=CURRENT_WEB_CFG.get("excel_path", "Template Column Excel Summary User.xlsx"),
        help="ตำแหน่งไฟล์ Excel ที่เชื่อมกับ SharePoint / OneDrive"
    )
    
    webhook_url = st.text_input(
        "⚡ Power Automate Webhook URL",
        value=CURRENT_WEB_CFG.get("webhook_url", ""),
        type="default",
        help="HTTP Webhook Trigger URL จาก Power Automate"
    )

    if st.button("💾 บันทึกการตั้งค่า (Save Settings)", use_container_width=True):
        save_web_config(gemini_key, excel_path, webhook_url)
        st.success("💾 บันทึกค่าการใช้งานล่าสุดเรียบร้อยแล้ว!")

    st.markdown("---")
    st.markdown("💡 **คู่มือใช้งาน**: วาง API Key ของคุณในช่องด้านบน ระบบจะจดจำคีย์ล่าสุดที่ใช้ไว้ตลอดเวลา จากนั้นแคปภาพหน้าจอและกดวางภาพได้เลยครับ")

# Auto-save key whenever used
if gemini_key and gemini_key != CURRENT_WEB_CFG.get("gemini_api_key", ""):
    save_web_config(gemini_key, excel_path, webhook_url)

# ==========================================
# 🚀 Main Page & Tabs Setup
# ==========================================
st.title("🔐 User Access Automation Web App")
st.caption("ระบบอ่านข้อมูลจากภาพแคปเจอร์ด้วย AI (Gemini Vision) ➔ บันทึก Excel ➔ ส่ง Email ตอบกลับอัตโนมัติ ➔ จัดการ Master Data")

master_data = load_master_data()

tab_ocr, tab_settings = st.tabs([
    "📄 1. สกัดข้อมูลภาพ & บันทึกผู้ใช้ (OCR & Create User)",
    "⚙️ 2. จัดการ Master Data & Email Templates"
])

# ==========================================
# 📄 TAB 1: OCR & Create User
# ==========================================
with tab_ocr:
    # --- Step 1: Upload Image & Application Selection ---
    st.subheader("📸 1. วางรูปภาพ (Clipboard Paste) หรืออัปโหลดไฟล์ภาพแคปเจอร์หน้าจอ")

    col_app, col_paste, col_up = st.columns([1, 1, 1.5])

    with col_app:
        target_app = st.selectbox(
            "📱 เลือกระบบ (Application)",
            ["🔍 Auto-Detect (อัตโนมัติ)", "VSM", "E-Travelling", "Forma", "Red plate", "Pandora"]
        )

    with col_paste:
        st.markdown("**📋 วางรูปภาพจาก Clipboard**")
        try:
            from streamlit_paste_button import paste_image_button
            paste_result = paste_image_button(
                label="📋 คลิกวางภาพจาก Clipboard (Paste)",
                background_color="#0d6efd",
                hover_background_color="#0b5ed7",
                text_color="#ffffff",
                errors="ignore"
            )
        except Exception:
            paste_result = None

    with col_up:
        uploaded_file = st.file_uploader(
            "📁 หรือเลือก/ลากวางไฟล์ภาพ (PNG, JPG, WEBP)",
            type=["png", "jpg", "jpeg", "webp"]
        )

    image = None
    if paste_result is not None and paste_result.image_data is not None:
        image = paste_result.image_data
    elif uploaded_file is not None:
        image = Image.open(uploaded_file)

    if image is not None:
        st.image(image, caption="ภาพแคปเจอร์ที่เลือก/วาง", use_container_width=True)

        if st.button("🤖 2. ประมวลผลและดึงข้อมูลจากภาพ (Extract Data)", type="primary"):
            with st.spinner("⏳ กำลังวิเคราะห์ภาพด้วย Gemini 2.0 Flash Vision AI..."):
                extracted, err_msg = extract_with_gemini_vision(image, gemini_key)
                if extracted:
                    st.session_state["extracted"] = extracted
                    st.success("✨ อ่านข้อมูลจากภาพด้วย Vision AI สำเร็จเป๊ะ 100%!")
                else:
                    st.error(f"⚠️ {err_msg}")

    # --- Step 2: Review & Edit Data Form ---
    if "extracted" in st.session_state:
        raw_data = st.session_state["extracted"]
        st.markdown("---")
        st.subheader("📝 3. ตรวจสอบและแก้ไขข้อมูลก่อนบันทึก (Review & Edit Data)")

        app_val = raw_data.get("Application", "E-Travelling")
        uid_val = raw_data.get("App user ID", "")
        neng_val = raw_data.get("Full Name Eng", "")
        nth_val = raw_data.get("Full Name Thai", "")
        email_val = raw_data.get("Email", "")
        pos_val = raw_data.get("Position", "")
        comp_val = raw_data.get("Company", "")
        branch_val = raw_data.get("Branch", "")

        current_companies = [c["Company"] for c in master_data.get("companies", []) if c.get("Company")]
        current_branches = master_data.get("branches", [])
        current_operators = master_data.get("operators", ["nawarutte.non@i24.co.th", "pawitporn.sae@i24.co.th"])

        col1, col2 = st.columns(2)
        with col1:
            edit_app = st.selectbox("Application", ["VSM", "E-Travelling", "Forma", "Red plate", "Pandora"], index=["VSM", "E-Travelling", "Forma", "Red plate", "Pandora"].index(app_val) if app_val in ["VSM", "E-Travelling", "Forma", "Red plate", "Pandora"] else 0)
            edit_neng = st.text_input("Full Name Eng (ชื่ออังกฤษ)", value=neng_val)
            edit_email = st.text_input("Email", value=email_val)
            
            comp_options = [""] + current_companies
            if comp_val and comp_val not in comp_options:
                comp_options.append(comp_val)
            edit_company = st.selectbox("Company (เลือกบริษัท)", comp_options, index=comp_options.index(comp_val) if comp_val in comp_options else 0)

        with col2:
            edit_uid = st.text_input("App user ID (รหัสผู้ใช้)", value=uid_val)
            edit_nth = st.text_input("Full Name Thai (ชื่อไทย)", value=nth_val)
            edit_pos = st.text_input("Position (ตำแหน่ง/กลุ่มผู้ใช้)", value=pos_val)
            
            branch_options = [""] + current_branches
            if branch_val and branch_val not in branch_options:
                branch_options.append(branch_val)
            edit_branch = st.selectbox("Branch (เลือกสาขา)", branch_options, index=branch_options.index(branch_val) if branch_val in branch_options else 0)

        st.markdown("#### 🔑 ข้อมูล Username & Password สำหรับส่ง Email")
        col_un, col_pw = st.columns(2)
        
        default_uname = edit_uid
        default_pwd = get_default_password_for_app(edit_app, edit_uid, edit_email, extracted_pwd=raw_data.get("Password", ""))

        with col_un:
            edit_email_username = st.text_input("🔑 Username (ส่ง Email)", value=default_uname)
        with col_pw:
            edit_email_password = st.text_input("🔐 Password (ส่ง Email)", value=default_pwd)

        # --- Step 3: Real-Time Live Email Preview ---
        st.markdown("---")
        st.subheader("📧 4. ตัวอย่าง Email ที่จะส่งหา User (Real-Time Live Preview)")

        col_sender, _ = st.columns([1, 1])
        with col_sender:
            edit_operator = st.selectbox(
                "👤 เลือกบัญชีผู้ส่ง Email (Sender / Operator)",
                current_operators,
                index=0,
                help="ระบบจะดึง Template อีเมลเฉพาะของบุคคลนี้มาแสดงและส่งออกจากกล่องข้อความของคนนี้"
            )

        active_tpl = get_operator_template(master_data, edit_operator, edit_app)
        email_subject = active_tpl.get("subject", f"ข้อมูลการเข้าระบบ {edit_app}")
        
        link_str = ""
        if edit_app == "VSM":
            link_str = "https://vsm.mgc-asia.com/Pages/Home.aspx"
        elif edit_app == "E-Travelling":
            link_str = "http://travelling.mgc-asia.com/"
        elif edit_app == "Red plate":
            link_str = "https://redplate-frontend.azurewebsites.net/signin/?redirect_url=%2Freport%2F%3Ftype%3Dred-plate-transaction"

        plain_body_preview, _ = build_email_body(edit_app, edit_email_username, edit_email_password, link_str, custom_template=active_tpl)

        email_preview_text = f"""========================================
📧 TEMPLATE สำหรับตอบ EMAIL ({edit_app}) - โดย {edit_operator}
========================================
To:       {edit_email if edit_email else '(ยังไม่ได้ระบุ Email)'}
Subject:  {email_subject}
----------------------------------------
{plain_body_preview}
========================================"""

        st.code(email_preview_text, language="text")

        # --- Step 4: Confirm Action Buttons ---
        st.markdown("---")
        st.subheader("🚀 5. เลือกคำสั่งบันทึกข้อมูล (Save & Send Options)")

        col_act1, col_act2 = st.columns(2)
        btn_excel_only = col_act1.button("📊 1. บันทึกลง Excel อย่างเดียว (ไม่ส่ง Email)", use_container_width=True)
        btn_excel_email = col_act2.button("📧 2. บันทึกลง Excel และส่ง Email", type="primary", use_container_width=True)

        if btn_excel_only or btn_excel_email:
            should_send_email = True if btn_excel_email else False
            reviewed_data = {
                "Application": edit_app,
                "App user ID": edit_email_username,
                "Full Name Eng": edit_neng,
                "Full Name Thai": edit_nth,
                "Email": edit_email,
                "Position": edit_pos,
                "Company": edit_company,
                "Branch": edit_branch
            }
            
            mapped_data = process_mapping(reviewed_data, master_companies=master_data.get("companies", []))
            export_to_excel(mapped_data, excel_path)

            if webhook_url:
                trigger_power_automate_webhook(
                    webhook_url, reviewed_data, mapped_data,
                    custom_username=edit_email_username,
                    custom_password=edit_email_password,
                    send_email=should_send_email,
                    operator=edit_operator,
                    custom_template=active_tpl
                )
            
            st.success("✅ ดำเนินการสำเร็จ!")
            st.balloons()

# ==========================================
# ⚙️ TAB 2: Master Data & Email Templates
# ==========================================
with tab_settings:
    st.subheader("⚙️ จัดการ Master Data & Email Templates")
    st.caption("แก้ไขข้อมูลบริษัท, ตัวย่อ BU, รายชื่อสาขา และปรับแต่งข้อความอีเมลแยกตามรายบุคคล ข้อมูลทั้งหมดจะถูกจดจำอัตโนมัติ")

    subtab_comp, subtab_branch, subtab_ops, subtab_tpl, subtab_backup = st.tabs([
        "🏢 1. บริษัท & BU (Company & BU)",
        "📍 2. รายชื่อสาขา (Branches)",
        "👤 3. รายชื่อผู้ส่ง (Operators)",
        "📧 4. เทมเพลตอีเมล (Email Templates)",
        "💾 5. สำรองและคืนค่า (Backup & Restore)"
    ])

    # ------------------------------------------
    # Sub-tab 1: Company & BU (Management Table + Live Search + Action Badges + Micro-interactions)
    # ------------------------------------------
    with subtab_comp:
        st.markdown("### 🏢 จัดการรายชื่อบริษัท และตัวย่อ BU")
        st.caption("ระบบล็อก **1 บริษัท = 1 BU เสมอ** สามารถค้นหา กรอง และจัดการข้อมูลได้แบบเรียลไทม์")

        companies_list = master_data.get("companies", [])

        # --- Top Search & Sort Bar ---
        col_search, col_sort = st.columns([2.5, 1.2])
        with col_search:
            search_comp_txt = st.text_input(
                "🔍 ค้นหาชื่อบริษัท หรือตัวย่อ BU...",
                placeholder="พิมพ์คำค้นหา เช่น Honda, MCR, BMW...",
                key="comp_search_input"
            )
        with col_sort:
            sort_comp_by = st.selectbox(
                "🔃 เรียงตาม",
                ["ชื่อบริษัท (A-Z)", "ชื่อบริษัท (Z-A)", "ตัวย่อ BU (A-Z)", "ตัวย่อ BU (Z-A)"],
                key="comp_sort_select"
            )

        # Filtering
        filtered_comps = companies_list
        if search_comp_txt.strip():
            q = search_comp_txt.strip().lower()
            filtered_comps = [c for c in companies_list if q in c["Company"].lower() or q in c["BU"].lower()]

        # Sorting
        if sort_comp_by == "ชื่อบริษัท (A-Z)":
            filtered_comps = sorted(filtered_comps, key=lambda x: x["Company"].lower())
        elif sort_comp_by == "ชื่อบริษัท (Z-A)":
            filtered_comps = sorted(filtered_comps, key=lambda x: x["Company"].lower(), reverse=True)
        elif sort_comp_by == "ตัวย่อ BU (A-Z)":
            filtered_comps = sorted(filtered_comps, key=lambda x: x["BU"].lower())
        elif sort_comp_by == "ตัวย่อ BU (Z-A)":
            filtered_comps = sorted(filtered_comps, key=lambda x: x["BU"].lower(), reverse=True)

        st.caption(f"📊 แสดง **{len(filtered_comps)}** จากทั้งหมด **{len(companies_list)}** บริษัท")

        # --- Add / Edit Company Drawer with Real-Time Validation ---
        with st.expander("➕ เพิ่มบริษัทใหม่ / แก้ไขตัวย่อ BU", expanded=False):
            col_in_c1, col_in_c2 = st.columns([2, 1])
            with col_in_c1:
                input_comp_name = st.text_input("ชื่อบริษัท (Company Name)*", placeholder="เช่น Test Automobile Co., Ltd.", key="input_comp_name_field")
            with col_in_c2:
                input_bu_name = st.text_input("ตัวย่อ BU*", placeholder="เช่น TAB", key="input_bu_name_field")

            # Real-Time Live Validation
            if input_comp_name.strip():
                match_existing = next((c for c in companies_list if c["Company"].strip().lower() == input_comp_name.strip().lower()), None)
                if match_existing:
                    st.info(f"💡 **พบข้อมูลเดิม**: ปัจจุบัน BU คือ `{match_existing['BU']}` — หากกดบันทึกจะเป็นการ **อัปเดตตัวย่อ BU**")
                else:
                    st.success(f"✅ **ชื่อบริษัทใหม่**: พร้อมเพิ่มเข้าสู่ระบบ (1 บริษัท = 1 BU)")

            if st.button("💾 บันทึกบริษัท & BU", type="primary", key="btn_save_comp_action"):
                if not input_comp_name.strip() or not input_bu_name.strip():
                    st.error("⚠️ กรุณากรอกทั้งชื่อบริษัทและตัวย่อ BU ให้ครบถ้วน")
                else:
                    c_clean = input_comp_name.strip()
                    bu_clean = input_bu_name.strip()
                    found = False
                    for item in companies_list:
                        if item["Company"].strip().lower() == c_clean.lower():
                            item["BU"] = bu_clean
                            found = True
                            break
                    if not found:
                        companies_list.append({"Company": c_clean, "BU": bu_clean})
                    
                    master_data["companies"] = companies_list
                    if save_master_data(master_data):
                        st.toast(f"✅ บันทึกบริษัท '{c_clean}' (BU: {bu_clean}) สำเร็จ!", icon="🏢")
                        st.rerun()

        # --- Management Table with Row-Level Badges & Actions ---
        st.markdown("---")
        
        # Table Header
        h_col1, h_col2, h_col3, h_col4 = st.columns([0.6, 3.8, 1.4, 1.2])
        h_col1.markdown("**#**")
        h_col2.markdown("**ชื่อบริษัท (Company Name)**")
        h_col3.markdown("**🏷️ ตัวย่อ BU**")
        h_col4.markdown("**การจัดการ (Action)**")

        st.markdown("<hr style='margin-top:0;margin-bottom:8px;'>", unsafe_allow_html=True)

        if not filtered_comps:
            st.info("🔍 ไม่พบบริษัทที่ตรงกับคำค้นหา")
        else:
            for idx, c_entry in enumerate(filtered_comps):
                row_c1, row_c2, row_c3, row_c4 = st.columns([0.6, 3.8, 1.4, 1.2])
                row_c1.write(f"**{idx + 1}**")
                row_c2.write(f"🏢 {c_entry['Company']}")
                row_c3.markdown(f"`🏷️ {c_entry['BU']}`")
                
                # Action Button: Delete
                if row_c4.button("🗑️ ลบ", key=f"del_comp_btn_{c_entry['Company']}", help=f"ลบบริษัท {c_entry['Company']}"):
                    master_data["companies"] = [item for item in companies_list if item["Company"] != c_entry["Company"]]
                    if save_master_data(master_data):
                        st.toast(f"🗑️ ลบบริษัท '{c_entry['Company']}' เรียบร้อยแล้ว", icon="🗑️")
                        st.rerun()

    # ------------------------------------------
    # Sub-tab 2: Branches (Interactive Tag Chips + Real-Time Validation + Toast)
    # ------------------------------------------
    with subtab_branch:
        st.markdown("### 📍 จัดการรายชื่อสาขา (Branches)")
        st.caption("คลิกปุ่ม ✕ บนป้ายสาขาเพื่อลบออกทันที หรือพิมพ์ชื่อสาขาใหม่เพื่อเพิ่มลงในระบบ")

        branches_list = master_data.get("branches", [])

        # --- Quick Add Branch with Real-Time Validation ---
        col_b_in, col_b_btn = st.columns([3, 1])
        with col_b_in:
            new_branch_input = st.text_input(
                "➕ เพิ่มสาขาใหม่",
                placeholder="พิมพ์ชื่อสาขา เช่น ระยอง, ขอนแก่น, สาทร...",
                key="new_branch_quick_input"
            )
            # Live Validation
            if new_branch_input.strip():
                if new_branch_input.strip() in branches_list:
                    st.warning(f"⚠️ มีสาขา '{new_branch_input.strip()}' อยู่ในระบบแล้ว")
                else:
                    st.success(f"✅ สาขา '{new_branch_input.strip()}' พร้อมเพิ่มลงในระบบ")

        with col_b_btn:
            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
            if st.button("➕ เพิ่มสาขา", type="primary", use_container_width=True, key="btn_add_branch_chip"):
                b_clean = new_branch_input.strip()
                if not b_clean:
                    st.error("กรุณากรอกชื่อสาขา")
                elif b_clean in branches_list:
                    st.warning("มีชื่อสาขานี้อยู่ในระบบแล้ว")
                else:
                    branches_list.append(b_clean)
                    master_data["branches"] = branches_list
                    if save_master_data(master_data):
                        st.toast(f"📍 เพิ่มสาขา '{b_clean}' สำเร็จ!", icon="📍")
                        st.rerun()

        # --- Search / Filter for branches ---
        st.markdown("---")
        col_b_search, col_b_count = st.columns([2.5, 1.5])
        with col_b_search:
            search_branch_txt = st.text_input(
                "🔍 กรองค้นหาสาขา...",
                placeholder="พิมพ์ชื่อสาขาเพื่อกรอง...",
                key="search_branch_chip_input"
            )
        with col_b_count:
            st.markdown("<div style='height: 32px;'></div>", unsafe_allow_html=True)
            st.caption(f"📊 ทั้งหมด **{len(branches_list)}** สาขา")

        filtered_branches = branches_list
        if search_branch_txt.strip():
            filtered_branches = [b for b in branches_list if search_branch_txt.strip().lower() in b.lower()]

        # --- Interactive Tag Chips Grid ---
        st.markdown("##### 🏷️ รายชื่อสาขา (คลิก ✕ เพื่อลบ):")
        
        # Display as responsive grid of chip buttons
        chip_cols = st.columns(4)
        for idx, br_name in enumerate(filtered_branches):
            col_target = chip_cols[idx % 4]
            with col_target:
                if col_target.button(f"📍 {br_name} ✕", key=f"chip_del_{br_name}", use_container_width=True, help=f"คลิกเพื่อลบสาขา {br_name}"):
                    branches_list.remove(br_name)
                    master_data["branches"] = branches_list
                    if save_master_data(master_data):
                        st.toast(f"🗑️ ลบสาขา '{br_name}' เรียบร้อยแล้ว", icon="🗑️")
                        st.rerun()

    # ------------------------------------------
    # Sub-tab 3: Operators / Senders
    # ------------------------------------------
    with subtab_ops:
        st.markdown("### 👤 จัดการรายชื่อผู้ส่ง Email (Operators)")
        
        ops_list = master_data.get("operators", [])
        st.markdown("##### 📋 บัญชีผู้ส่งที่ใช้งานได้ปัจจุบัน:")
        for op in ops_list:
            st.markdown(f"- ✉️ `{op}`")

        col_add_op, col_del_op = st.columns(2)
        with col_add_op:
            st.markdown("##### ➕ เพิ่มบัญชีผู้ส่ง")
            with st.form("form_add_operator"):
                new_op_email = st.text_input("อีเมลผู้ส่ง (Office 365)*", placeholder="เช่น name.sur@i24.co.th")
                btn_save_op = st.form_submit_button("💾 เพิ่มผู้ส่ง")
                
                if btn_save_op:
                    if not new_op_email.strip() or "@" not in new_op_email:
                        st.error("กรุณากรอกอีเมลที่ถูกต้อง")
                    elif new_op_email.strip() in ops_list:
                        st.warning("มีอีเมลนี้อยู่ในระบบแล้ว")
                    else:
                        ops_list.append(new_op_email.strip())
                        master_data["operators"] = ops_list
                        # Copy default templates for new operator
                        if "templates" not in master_data:
                            master_data["templates"] = {}
                        if new_op_email.strip() not in master_data["templates"]:
                            master_data["templates"][new_op_email.strip()] = json.loads(json.dumps(DEFAULT_MASTER_DATA["templates"]["nawarutte.non@i24.co.th"]))
                        
                        if save_master_data(master_data):
                            st.toast(f"👤 เพิ่มผู้ส่ง '{new_op_email.strip()}' สำเร็จ!", icon="👤")
                            st.rerun()

        with col_del_op:
            st.markdown("##### 🗑️ ลบบัญชีผู้ส่ง")
            with st.form("form_del_operator"):
                del_op_email = st.selectbox("เลือกบัญชีที่ต้องการลบ", ops_list if ops_list else [""])
                btn_del_op = st.form_submit_button("🗑️ ยืนยันลบผู้ส่ง")
                
                if btn_del_op:
                    if len(ops_list) <= 1:
                        st.error("ต้องมีผู้ส่งในระบบอย่างน้อย 1 คน")
                    elif del_op_email and del_op_email in ops_list:
                        ops_list.remove(del_op_email)
                        master_data["operators"] = ops_list
                        if del_op_email in master_data.get("templates", {}):
                            del master_data["templates"][del_op_email]
                        if save_master_data(master_data):
                            st.toast(f"🗑️ ลบผู้ส่ง '{del_op_email}' สำเร็จ!", icon="🗑️")
                            st.rerun()

    # ------------------------------------------
    # Sub-tab 4: Personalized Email Template Editor
    # ------------------------------------------
    with subtab_tpl:
        st.markdown("### 📧 ตัวแก้ไข Template Email แยกตาม User & Application")
        st.caption("ปรับแต่งหัวข้อ คำขึ้นต้น และสำนวนภาษาให้ตรงตามความถนัดของแต่ละคน ระบบจะจำค่าแยกเป็นรายบุคคล")

        ops_list = master_data.get("operators", ["nawarutte.non@i24.co.th", "pawitporn.sae@i24.co.th"])
        app_list = ["VSM", "E-Travelling", "Forma", "Red plate", "Pandora"]

        col_sel_user, col_sel_app = st.columns(2)
        with col_sel_user:
            edit_tpl_user = st.selectbox("👤 1. เลือก User ผู้ส่ง", ops_list, index=0, key="tpl_user_sel")
        with col_sel_app:
            edit_tpl_app = st.selectbox("📱 2. เลือกระบบ (Application)", app_list, index=0, key="tpl_app_sel")

        current_tpl = get_operator_template(master_data, edit_tpl_user, edit_tpl_app)

        with st.form("form_edit_email_template"):
            st.markdown(f"#### ✍️ แก้ไข Template ของ `{edit_tpl_user}` สำหรับระบบ `{edit_tpl_app}`")
            
            in_subject = st.text_input("📝 Subject (หัวข้ออีเมล)*", value=current_tpl.get("subject", f"ข้อมูลการเข้าระบบ {edit_tpl_app}"))
            in_greeting = st.text_input("👋 คำขึ้นต้น / คำทักทาย*", value=current_tpl.get("greeting", "เรียน ผู้ใช้งานระบบ"))
            in_intro = st.text_area("💬 ข้อความเกริ่นนำ (Intro Text)*", value=current_tpl.get("intro", f"ให้เข้าใช้งานโดย User & Password ตามด้านล่างนี้ครับ"), height=110)
            
            st.info("💡 หมายเหตุ: บรรทัด User, Password และ Link จะถูกสร้างและจัดรูปแบบให้อัตโนมัติตามประเภทโปรแกรม")

            col_btn_tpl1, col_btn_tpl2 = st.columns(2)
            btn_save_tpl = col_btn_tpl1.form_submit_button("💾 บันทึก Template นี้", type="primary", use_container_width=True)
            btn_reset_tpl = col_btn_tpl2.form_submit_button("🔄 คืนค่าเริ่มต้นของ App นี้", use_container_width=True)

            if btn_save_tpl:
                if "templates" not in master_data:
                    master_data["templates"] = {}
                if edit_tpl_user not in master_data["templates"]:
                    master_data["templates"][edit_tpl_user] = {}
                
                master_data["templates"][edit_tpl_user][edit_tpl_app] = {
                    "subject": in_subject.strip(),
                    "greeting": in_greeting.strip(),
                    "intro": in_intro.strip()
                }
                if save_master_data(master_data):
                    st.toast(f"📧 บันทึก Template '{edit_tpl_app}' ของ '{edit_tpl_user}' สำเร็จ!", icon="💾")
                    st.rerun()

            if btn_reset_tpl:
                def_tpl = DEFAULT_MASTER_DATA["templates"]["nawarutte.non@i24.co.th"].get(edit_tpl_app, {})
                if "templates" in master_data and edit_tpl_user in master_data["templates"]:
                    master_data["templates"][edit_tpl_user][edit_tpl_app] = def_tpl
                    if save_master_data(master_data):
                        st.toast(f"🔄 คืนค่าเริ่มต้น Template '{edit_tpl_app}' สำเร็จ!", icon="🔄")
                        st.rerun()

        # Real-Time Live Preview of the edited template
        st.markdown("---")
        st.markdown("##### 👁️ ตัวอย่างผลลัพธ์ของ Template นี้ (Live Sample Preview):")
        sample_link = "https://vsm.mgc-asia.com/Pages/Home.aspx" if edit_tpl_app == "VSM" else ("http://travelling.mgc-asia.com/" if edit_tpl_app == "E-Travelling" else ("https://redplate-frontend.azurewebsites.net/..." if edit_tpl_app == "Red plate" else ""))
        sample_plain, sample_html = build_email_body(edit_tpl_app, "somchai.pra", "p@ssw0rdsom", sample_link, custom_template={"greeting": in_greeting, "intro": in_intro})
        st.code(f"""To:       user.example@mgc-asia.com
Subject:  {in_subject}
----------------------------------------
{sample_plain}""", language="text")

    # ------------------------------------------
    # Sub-tab 5: Backup & Restore Master Data
    # ------------------------------------------
    with subtab_backup:
        st.markdown("### 💾 สำรองและกู้คืน Master Data (Backup & Restore)")
        
        json_str = json.dumps(master_data, ensure_ascii=False, indent=2)
        st.download_button(
            label="📥 ดาวน์โหลดไฟล์สำรอง (Export master_data.json)",
            data=json_str,
            file_name="master_data.json",
            mime="application/json",
            use_container_width=True
        )

        st.markdown("---")
        st.markdown("##### 📤 กู้คืน / นำเข้าไฟล์ Master Data (Import JSON):")
        uploaded_json = st.file_uploader("เลือกไฟล์ master_data.json ที่สำรองไว้", type=["json"])
        if uploaded_json is not None:
            try:
                imported_data = json.load(uploaded_json)
                if st.button("🔄 ยืนยันนำเข้าข้อมูลนี้ (Apply Imported Data)", type="primary"):
                    if save_master_data(imported_data):
                        st.toast("✅ นำเข้าข้อมูล Master Data สำเร็จ!", icon="📥")
                        st.rerun()
            except Exception as ex:
                st.error(f"ไฟล์ JSON ไม่ถูกต้อง: {ex}")

        st.markdown("---")
        if st.button("⚠️ คืนค่า Master Data ทั้งหมดเป็นค่าเริ่มต้นจากโรงงาน (Reset All to Factory Defaults)"):
            if save_master_data(DEFAULT_MASTER_DATA):
                st.toast("🔄 คืนค่า Master Data ทั้งหมดเป็นค่าเริ่มต้นเรียบร้อยแล้ว!", icon="🔄")
                st.rerun()

